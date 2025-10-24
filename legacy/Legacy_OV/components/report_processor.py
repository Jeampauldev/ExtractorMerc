"""
Procesador avanzado de reportes descargados con capacidades de:
- Detección automática de tipos de archivo
- Clasificación por empresa y tipo de reporte
- Validación de contenido
- Organización automática en directorios
- Detección de duplicados
- Generación de reportes de procesamiento
- Manejo de errores robusto
- Logging detallado
- Configuración flexible de reglas de procesamiento
"""

import logging
import os
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Union, Any, Tuple, Callable
from dataclasses import dataclass
from enum import Enum
import re
import hashlib
import json
from datetime import datetime, date
import mimetypes
import csv
import openpyxl
from openpyxl import load_workbook

# Configuración del logger
logger = logging.getLogger('REPORT-PROCESSOR')

class FileType(Enum):
    """Tipos de archivo soportados"""
    EXCEL = "excel"
    CSV = "csv"
    PDF = "pdf"
    TXT = "txt"
    JSON = "json"
    XML = "xml"
    ZIP = "zip"
    RAR = "rar"
    DOC = "doc"
    DOCX = "docx"
    UNKNOWN = "unknown"

class ProcessingStatus(Enum):
    """Estados de procesamiento"""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    SKIPPED = "skipped"
    DUPLICATE = "duplicate"

@dataclass
class FileMetadata:
    """Metadatos de archivo procesado"""
    original_path: str
    final_path: str
    file_type: FileType
    size_bytes: int
    created_at: datetime
    processed_at: datetime
    checksum: str
    company: str = ""
    report_type: str = ""
    date_range: str = ""
    row_count: int = 0
    column_count: int = 0
    is_valid: bool = True
    error_message: str = ""

@dataclass
class ProcessingRule:
    """Regla de procesamiento de archivos"""
    name: str
    pattern: str # Regex pattern para nombre de archivo
    company: str
    report_type: str
    target_directory: str
    rename_template: str # Template para renombrado: {company}_{type}_{date}.{ext}
    processor_function: Optional[Callable] = None
    validation_rules: List[str] = None

    def __post_init__(self):
        if self.validation_rules is None:
            self.validation_rules = []

class ReportProcessor:
    """
    Procesador avanzado de reportes descargados
    """

    def __init__(self, base_directory: str = "downloads", use_defaults: bool = True):
        """
        Inicializar el procesador de reportes

        Args:
            base_directory: Directorio base para procesamiento
            use_defaults: Si usar reglas de procesamiento por defecto
        """
        self.base_directory = Path(base_directory)
        self.logger = logger # Usar el logger configurado globalmente

        # Crear directorio base si no existe
        self.base_directory.mkdir(parents=True, exist_ok=True)

        # Configurar subdirectorios
        self.processed_dir = self.base_directory / "processed"
        self.failed_dir = self.base_directory / "failed"
        self.duplicates_dir = self.base_directory / "duplicates"

        for directory in [self.processed_dir, self.failed_dir, self.duplicates_dir]:
            directory.mkdir(parents=True, exist_ok=True)

        # Inicializar estructuras de datos
        self.processing_rules = []

        # Archivos procesados y checksums para detección de duplicados
        self.processed_files = []
        self.file_checksums = {} # Para detección de duplicados

        # Estadísticas de procesamiento
        self.processing_stats = {
            'total_processed': 0,
            'successful': 0,
            'failed': 0,
            'duplicates': 0,
            'total_size_mb': 0.0
        }

        # Cargar reglas por defecto si se solicita
        if use_defaults:
            self._setup_default_rules()

    def _setup_default_rules(self):
        """Configurar reglas de procesamiento por defecto"""
        try:
            from src.config.default_configs import get_default_processing_rules
            for rule in get_default_processing_rules():
                self.add_processing_rule(rule)
        except ImportError:
            self.logger.warning("No se pudieron cargar las reglas por defecto")
            return

    def add_processing_rule(self, rule: ProcessingRule) -> bool:
        """
        Agregar una regla de procesamiento

        Args:
            rule: Regla de procesamiento a agregar

        Returns:
            bool: True si se agregó exitosamente
        """
        try:
            # Validar que el patrón regex sea válido
            re.compile(rule.pattern)

            self.processing_rules.append(rule)
            self.logger.info(f"Regla de procesamiento agregada: {rule.name}")
            return True

        except re.error as e:
            self.logger.error(f"Pattern regex inválido en regla {rule.name}: {e}")
            return False
        except Exception as e:
            self.logger.error(f"Error agregando regla {rule.name}: {e}")
            return False

    def process_file(self, file_path: Union[str, Path], 
                    company: str = None, 
                    report_type: str = None) -> FileMetadata:
        """
        Procesar un archivo individual

        Args:
            file_path: Ruta del archivo a procesar
            company: Empresa (opcional, se detecta automáticamente)
            report_type: Tipo de reporte (opcional, se detecta automáticamente)

        Returns:
            FileMetadata: Metadatos del archivo procesado
        """
        file_path = Path(file_path)
        start_time = datetime.now()

        self.logger.info(f"Procesando archivo: {file_path}")

        try:
            # Validaciones básicas
            if not file_path.exists():
                raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

            # Obtener información básica del archivo
            file_size = file_path.stat().st_size
            file_type = self._detect_file_type(file_path)
            checksum = self._calculate_checksum(file_path)

            # Verificar si es duplicado
            if self._is_duplicate(checksum, file_path):
                return self._handle_duplicate_file(file_path, checksum)

            # Encontrar regla de procesamiento aplicable
            rule = self._find_matching_rule(file_path.name)

            # Detectar empresa y tipo de reporte
            detected_company = company or (rule.company if rule else self._detect_company(file_path))
            detected_type = report_type or (rule.report_type if rule else self._detect_report_type(file_path))

            # Validar archivo según reglas
            validation_result = self._validate_file(file_path, rule)
            if not validation_result['is_valid']:
                return self._handle_failed_file(file_path, validation_result['error'])

            # Determinar ruta de destino
            target_path = self._determine_target_path(file_path, detected_company, detected_type, rule)

            # Analizar contenido del archivo
            content_info = self._analyze_file_content(file_path, file_type)

            # Mover archivo a destino final
            final_path = self._move_file_to_target(file_path, target_path)

            # Crear metadatos
            metadata = FileMetadata(
                original_path=str(file_path),
                final_path=str(final_path),
                file_type=file_type,
                size_bytes=file_size,
                created_at=datetime.fromtimestamp(file_path.stat().st_ctime),
                processed_at=start_time,
                checksum=checksum,
                company=detected_company,
                report_type=detected_type,
                row_count=content_info.get('rows', 0),
                column_count=content_info.get('columns', 0),
                is_valid=True
            )

            # Registrar archivo procesado
            self.processed_files.append(metadata)
            self.file_checksums[checksum] = str(final_path)

            # Actualizar estadísticas
            self._update_stats(metadata, ProcessingStatus.COMPLETED)

            self.logger.info(f"Archivo procesado exitosamente: {final_path}")
            return metadata

        except Exception as e:
            self.logger.error(f"Error procesando archivo {file_path}: {e}")
            return self._handle_failed_file(file_path, str(e))

    def _detect_file_type(self, file_path: Path) -> FileType:
        """Detectar tipo de archivo basado en extensión"""
        extension = file_path.suffix.lower()

        type_mapping = {
            '.xlsx': FileType.EXCEL,
            '.xls': FileType.EXCEL,
            '.csv': FileType.CSV,
            '.pdf': FileType.PDF,
            '.txt': FileType.TXT,
            '.json': FileType.JSON,
            '.xml': FileType.XML,
            '.zip': FileType.ZIP,
            '.rar': FileType.RAR,
            '.doc': FileType.DOC,
            '.docx': FileType.DOCX
        }

        return type_mapping.get(extension, FileType.UNKNOWN)

    def _calculate_checksum(self, file_path: Path) -> str:
        """Calcular checksum MD5 del archivo"""
        hash_md5 = hashlib.md5()

        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)

        return hash_md5.hexdigest()

    def _is_duplicate(self, checksum: str, file_path: Path) -> bool:
        """Verificar si el archivo es duplicado"""
        if checksum in self.file_checksums:
            existing_path = self.file_checksums[checksum]
            if Path(existing_path).exists():
                self.logger.info(f"Archivo duplicado detectado: {file_path}")
                return True
            else:
                # El archivo original ya no existe, remover del registro
                del self.file_checksums[checksum]

        return False

    def _find_matching_rule(self, filename: str) -> Optional[ProcessingRule]:
        """Encontrar regla de procesamiento que coincida con el archivo"""
        for rule in self.processing_rules:
            if re.search(rule.pattern, filename, re.IGNORECASE):
                self.logger.debug(f"Regla encontrada: {rule.name} para {filename}")
                return rule
        return None

    def _detect_company(self, file_path: Path) -> str:
        """Detectar empresa basada en nombre de archivo y ruta"""
        filename = file_path.name.lower()
        path_str = str(file_path).lower()

        company_patterns = {
            'afinia': ['afinia', 'electrocosta'],
            'aire': ['aire', 'epm'],
            'enel': ['enel', 'codensa'],
            'emcali': ['emcali'],
            'epm': ['epm']
        }

        for company, patterns in company_patterns.items():
            if any(pattern in filename or pattern in path_str for pattern in patterns):
                return company

        return "general"

    def _detect_report_type(self, file_path: Path) -> str:
        """Detectar tipo de reporte basado en nombre de archivo"""
        filename = file_path.name.lower()

        type_patterns = {
            'pqr': ['pqr', 'peticion', 'queja', 'reclamo', 'solicitud'],
            'financial': ['factura', 'pago', 'cobro', 'financiero', 'recaudo'],
            'consumption': ['consumo', 'lectura', 'medidor'],
            'outages': ['interrupciones', 'cortes', 'saidi', 'saifi'],
            'commercial': ['comercial', 'cliente', 'contrato'],
            'technical': ['tecnico', 'mantenimiento', 'red'],
            'quality': ['calidad', 'indicadores', 'kpi']
        }

        for report_type, patterns in type_patterns.items():
            if any(pattern in filename for pattern in patterns):
                return report_type

        return "general"

    def _validate_file(self, file_path: Path, rule: Optional[ProcessingRule]) -> Dict[str, Any]:
        """Validar archivo según reglas definidas"""
        result = {'is_valid': True, 'error': ''}

        try:
            file_stat = file_path.stat()

            # Validación básica: archivo no vacío
            if file_stat.st_size == 0:
                result['is_valid'] = False
                result['error'] = "Archivo vacío"
                return result

            # Aplicar reglas de validación específicas si existen
            if rule and rule.validation_rules:
                for validation_rule in rule.validation_rules:
                    if not self._apply_validation_rule(file_path, validation_rule):
                        result['is_valid'] = False
                        result['error'] = f"Falló validación: {validation_rule}"
                        return result

            return result

        except Exception as e:
            result['is_valid'] = False
            result['error'] = str(e)
            return result

    def _apply_validation_rule(self, file_path: Path, rule: str) -> bool:
        """Aplicar una regla de validación específica"""
        try:
            if rule.startswith("min_size:"):
                min_size = int(rule.split(":")[1])
                return file_path.stat().st_size >= min_size

            elif rule.startswith("max_age_days:"):
                max_days = int(rule.split(":")[1])
                file_age = (datetime.now().timestamp() - file_path.stat().st_mtime) / 86400
                return file_age <= max_days

            elif rule.startswith("required_columns:"):
                columns = rule.split(":")[1].split(",")
                return self._validate_required_columns(file_path, columns)

            elif rule == "not_empty":
                return file_path.stat().st_size > 0

            else:
                self.logger.warning(f"Regla de validación no reconocida: {rule}")
                return True

        except Exception as e:
            self.logger.error(f"Error aplicando regla de validación {rule}: {e}")
            return False

    def _validate_required_columns(self, file_path: Path, required_columns: List[str]) -> bool:
        """Validar que el archivo contenga las columnas requeridas"""
        try:
            file_type = self._detect_file_type(file_path)

            if file_type == FileType.CSV:
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader, [])

            elif file_type == FileType.EXCEL:
                workbook = load_workbook(file_path)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]

            else:
                return True # No aplicable a otros tipos de archivo

            # Comparar headers (case insensitive)
            headers_lower = [h.lower().strip() if h else '' for h in headers]
            required_lower = [col.lower().strip() for col in required_columns]

            return all(req_col in headers_lower for req_col in required_lower)

        except Exception as e:
            self.logger.error(f"Error validando columnas requeridas: {e}")
            return False

    def _determine_target_path(self, file_path: Path, company: str, report_type: str, 
                              rule: Optional[ProcessingRule]) -> Path:
        """Determinar la ruta de destino para el archivo"""
        # Usar directorio de la regla si está definido
        if rule and rule.target_directory:
            # Permitir templates en el directorio de destino
            target_dir = rule.target_directory.format(
                company=company,
                type=report_type,
                date=date.today().strftime("%Y-%m")
            )
        else:
            target_dir = f"{company}/{report_type}"

        full_target_dir = self.processed_dir / target_dir
        full_target_dir.mkdir(parents=True, exist_ok=True)

        # Determinar nombre de archivo final
        if rule and rule.rename_template:
            new_name = self._generate_filename(file_path, rule.rename_template, company, report_type)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            extension = file_path.suffix
            new_name = f"{company}_{report_type}_{timestamp}{extension}"

        return full_target_dir / new_name

    def _generate_filename(self, file_path: Path, template: str, company: str, report_type: str) -> str:
        """Generar nombre de archivo usando template"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        date_str = date.today().strftime("%Y-%m-%d")
        extension = file_path.suffix.lstrip('.')

        return template.format(
            company=company,
            type=report_type,
            date=date_str,
            timestamp=timestamp,
            ext=extension,
            original=file_path.stem
        )

    def _analyze_file_content(self, file_path: Path, file_type: FileType) -> Dict[str, Any]:
        """Analizar contenido del archivo para obtener metadatos"""
        content_info = {'rows': 0, 'columns': 0, 'sheets': 0}

        try:
            if file_type == FileType.CSV:
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    content_info['rows'] = len(rows)
                    content_info['columns'] = len(rows[0]) if rows else 0

            elif file_type == FileType.EXCEL:
                workbook = load_workbook(file_path, read_only=True)
                content_info['sheets'] = len(workbook.worksheets)

                # Analizar primera hoja
                sheet = workbook.active
                content_info['rows'] = sheet.max_row
                content_info['columns'] = sheet.max_column

            elif file_type == FileType.PDF:
                # Para PDFs, solo registrar tamaño
                content_info['size_mb'] = file_path.stat().st_size / (1024 * 1024)

        except Exception as e:
            self.logger.warning(f"No se pudo analizar contenido de {file_path}: {e}")

        return content_info

    def _move_file_to_target(self, source_path: Path, target_path: Path) -> Path:
        """Mover archivo a la ubicación de destino"""
        try:
            # Asegurar que no haya conflictos de nombres
            counter = 1
            original_target = target_path

            while target_path.exists():
                stem = original_target.stem
                suffix = original_target.suffix
                target_path = original_target.parent / f"{stem}_{counter}{suffix}"
                counter += 1

            # Mover archivo
            shutil.move(str(source_path), str(target_path))
            self.logger.debug(f"Archivo movido: {source_path} -> {target_path}")

            return target_path

        except Exception as e:
            self.logger.error(f"Error moviendo archivo {source_path} -> {target_path}: {e}")
            raise

    def _handle_duplicate_file(self, file_path: Path, checksum: str) -> FileMetadata:
        """Manejar archivo duplicado"""
        duplicate_path = self.duplicates_dir / file_path.name

        # Mover a directorio de duplicados
        if file_path.exists():
            try:
                shutil.move(str(file_path), str(duplicate_path))
            except Exception as e:
                self.logger.error(f"Error moviendo duplicado: {e}")

        metadata = FileMetadata(
            original_path=str(file_path),
            final_path=str(duplicate_path),
            file_type=self._detect_file_type(file_path),
            size_bytes=file_path.stat().st_size if file_path.exists() else 0,
            created_at=datetime.now(),
            processed_at=datetime.now(),
            checksum=checksum,
            is_valid=False,
            error_message="Archivo duplicado"
        )

        self._update_stats(metadata, ProcessingStatus.DUPLICATE)
        return metadata

    def _handle_failed_file(self, file_path: Path, error: str) -> FileMetadata:
        """Manejar archivo que falló en el procesamiento"""
        failed_path = self.failed_dir / file_path.name

        # Mover a directorio de fallidos
        if file_path.exists():
            try:
                shutil.move(str(file_path), str(failed_path))
            except Exception as e:
                self.logger.error(f"Error moviendo archivo fallido: {e}")

        metadata = FileMetadata(
            original_path=str(file_path),
            final_path=str(failed_path),
            file_type=self._detect_file_type(file_path),
            size_bytes=file_path.stat().st_size if file_path.exists() else 0,
            created_at=datetime.now(),
            processed_at=datetime.now(),
            checksum="",
            is_valid=False,
            error_message=error
        )

        self._update_stats(metadata, ProcessingStatus.FAILED)
        return metadata

    def _should_skip_file(self, file_path: Path) -> bool:
        """Determinar si un archivo debe ser saltado"""
        filename = file_path.name.lower()

        # Patrones de archivos a saltar
        skip_patterns = [
            r'^\.', # Archivos ocultos
            r'~\$', # Archivos temporales de Office
            r'\.tmp$', # Archivos temporales
            r'\.temp$', # Archivos temporales
            r'thumbs\.db$', # Thumbnails de Windows
            r'\.ds_store$', # Archivos de macOS
            r'\.crdownload$', # Descargas incompletas de Chrome
            r'\.part$' # Descargas incompletas
        ]

        return any(re.search(pattern, filename) for pattern in skip_patterns)

    def _update_stats(self, metadata: FileMetadata, status: ProcessingStatus):
        """Actualizar estadísticas de procesamiento"""
        self.processing_stats['total_processed'] += 1

        if status == ProcessingStatus.COMPLETED:
            self.processing_stats['successful'] += 1
        elif status == ProcessingStatus.FAILED:
            self.processing_stats['failed'] += 1
        elif status == ProcessingStatus.DUPLICATE:
            self.processing_stats['duplicates'] += 1

        self.processing_stats['total_size_mb'] += metadata.size_bytes / (1024 * 1024)

    def get_processing_stats(self) -> Dict[str, Any]:
        """Obtener estadísticas de procesamiento"""
        success_rate = 0.0
        if self.processing_stats['total_processed'] > 0:
            success_rate = self.processing_stats['successful'] / self.processing_stats['total_processed']

        return {
            **self.processing_stats,
            'success_rate': success_rate,
            'rules_count': len(self.processing_rules),
            'processed_files_count': len(self.processed_files)
        }

    def generate_processing_report(self) -> Dict[str, Any]:
        """Generar reporte completo de procesamiento"""
        stats = self.get_processing_stats()

        # Estadísticas por empresa
        by_company = {}
        for file_meta in self.processed_files:
            company = file_meta.company
            if company not in by_company:
                by_company[company] = {'count': 0, 'size_mb': 0.0, 'types': {}}

            by_company[company]['count'] += 1
            by_company[company]['size_mb'] += file_meta.size_bytes / (1024 * 1024)

            report_type = file_meta.report_type
            if report_type not in by_company[company]['types']:
                by_company[company]['types'][report_type] = 0
            by_company[company]['types'][report_type] += 1

        # Estadísticas por tipo de archivo
        file_types = {}
        for file_meta in self.processed_files:
            file_type = file_meta.file_type.value
            file_types[file_type] = file_types.get(file_type, 0) + 1

        return {
            'summary': stats,
            'by_company': by_company,
            'file_types': file_types,
            'directories': {
                'processed': str(self.processed_dir),
                'failed': str(self.failed_dir),
                'duplicates': str(self.duplicates_dir)
            },
            'generated_at': datetime.now().isoformat()
        }

    def save_processing_log(self, output_path: Path = None):
        """Guardar log de procesamiento en archivo JSON"""
        if output_path is None:
            output_path = self.base_directory / f"processing_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        log_data = {
            'processing_report': self.generate_processing_report(),
            'processed_files': [
                {
                    'original_path': f.original_path,
                    'final_path': f.final_path,
                    'file_type': f.file_type.value,
                    'size_bytes': f.size_bytes,
                    'processed_at': f.processed_at.isoformat(),
                    'company': f.company,
                    'report_type': f.report_type,
                    'is_valid': f.is_valid,
                    'error_message': f.error_message
                }
                for f in self.processed_files
            ]
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(log_data, f, indent=2, ensure_ascii=False)

        self.logger.info(f"Log de procesamiento guardado en: {output_path}")

# Función de utilidad para crear reglas de procesamiento
def create_processing_rule(name: str, pattern: str, company: str, 
                          report_type: str, **kwargs) -> ProcessingRule:
    """
    Crear una regla de procesamiento con parámetros opcionales

    Args:
        name: Nombre de la regla
        pattern: Patrón regex para coincidencia de archivos
        company: Empresa asociada
        report_type: Tipo de reporte
        **kwargs: Parámetros adicionales (target_directory, rename_template, etc.)

    Returns:
        ProcessingRule: Regla de procesamiento configurada
    """
    return ProcessingRule(
        name=name,
        pattern=pattern,
        company=company,
        report_type=report_type,
        target_directory=kwargs.get('target_directory', ''),
        rename_template=kwargs.get('rename_template', ''),
        processor_function=kwargs.get('processor_function'),
        validation_rules=kwargs.get('validation_rules', [])
    )

if __name__ == "__main__":
    # Ejemplo de uso
    print("=== ReportProcessor - Pruebas ===")

    # Crear procesador
    processor = ReportProcessor("test_downloads")

    # Agregar regla personalizada
    custom_rule = create_processing_rule(
        name="test_rule",
        pattern=r"test.*\.xlsx$",
        company="test_company",
        report_type="test_reports"
    )
    processor.add_processing_rule(custom_rule)

    # Mostrar estadísticas
    stats = processor.get_processing_stats()
    print(f"Reglas configuradas: {stats['rules_count']}")
    print(f"Archivos procesados: {stats['processed_files_count']}")

    # Generar reporte
    report = processor.generate_processing_report()
    print(f"Reporte generado en: {report['generated_at']}")
    print(f"Directorios configurados: {report['directories']}")
